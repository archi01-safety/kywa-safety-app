import streamlit as st

# [1] 페이지 설정 (반드시 최상단 배치)
st.set_page_config(page_title="KYWA AI 위험성평가 시스템 (Dev)", layout="wide", page_icon="🚨")

# [2] 필수 라이브러리 임포트
import os
import ssl
import json
import requests
import io
import datetime
import base64
import re
import pandas as pd
import numpy as np
import cv2
import plotly.express as px
import urllib.parse
import zipfile
import xml.etree.ElementTree as ET
import google.genai as genai
import openpyxl

# PDF 관련 라이브러리
from reportlab.lib.pagesizes import A4, A3, landscape, portrait
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Frame, PageTemplate, NextPageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Retry 및 예외 처리 라이브러리
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception
from google.genai.errors import APIError

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image

# --- [환경 설정 및 보안 우회] ---
os.environ['PYTHONHTTPSVERIFY'] = '0'
ssl._create_default_https_context = ssl._create_unverified_context

# --- [상수 및 구글 서비스 설정] ---
DRIVE_FOLDER_ID = "13RYVnDB7rrqLQYzB5Wa9WdWr0CHjm_MW"
SPREADSHEET_ID = "1kL18jQn5t0UX8ECpVEm3RHLQAWu7lum8_Wb-EtxkU5Q"
SHEET_NAME = "설문지 응답 시트1"  # 실제 구글시트 탭 이름

drive_service = None
sheets_service = None

if "gcp_service_account" in st.secrets:
    try:
        creds_info = st.secrets["gcp_service_account"]
        if isinstance(creds_info, (dict, st.runtime.secrets.AttrDict)):
            creds_dict = dict(creds_info)
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
            SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']
            creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
            drive_service = build('drive', 'v3', credentials=creds)
            sheets_service = build('sheets', 'v4', credentials=creds)
    except Exception as e:
        st.error(f"⚠️ GCP 인증 설정 오류: {e}")

# Gemini API 클라이언트 초기화
client = None
model_name = "gemini-3.6-flash"
if "GEMINI_API_KEY" in st.secrets:
    client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])

# --- [Retry 판별 전용 검증 함수] ---
def is_transient_error(exception):
    """503 과부하 또는 429 Rate Limit 에러 시에만 재시도 수행"""
    if isinstance(exception, APIError):
        if exception.code in [503, 429] or "UNAVAILABLE" in str(exception) or "high demand" in str(exception):
            return True
    return False

# --- [Tenacity 기반 Gemini API 안전 호출 함수] ---
@retry(
    stop=stop_after_attempt(5),               # 최대 5회 시도
    wait=wait_exponential(min=2, max=20),      # 2초 -> 4초 -> 8초 -> 16초 지수 대기
    retry=retry_if_exception(is_transient_error),
    reraise=True                              # 최종 실패 시 예외를 던져 에러 텍스트 저장을 방지
)
def call_gemini_api_safe(client, model, contents, config=None):
    """503/429 발생 시 자동으로 지수 대기 재시도를 수행하는 안전 호출 함수"""
    if config:
        return client.models.generate_content(model=model, contents=contents, config=config)
    return client.models.generate_content(model=model, contents=contents)

# --- [유틸리티 함수들] ---

def get_image_bytes_from_link(url_or_id):
    if not url_or_id or not isinstance(url_or_id, str):
        return None
    url = url_or_id.strip()
    if not url.startswith("http"):
        return None

    drive_id_match = re.search(r'd/([a-zA-Z0-9_-]+)', url) or re.search(r'id=([a-zA-Z0-9_-]+)', url)
    if drive_id_match:
        file_id = drive_id_match.group(1)
        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
    else:
        download_url = url

    try:
        response = requests.get(download_url, timeout=7, headers={'User-Agent': 'Mozilla/5.0'})
        if response.status_code == 200:
            img_raw = Image.open(io.BytesIO(response.content))
            img_raw.thumbnail((300, 300), Image.Resampling.LANCZOS)
            buffer = io.BytesIO()
            img_raw.convert("RGB").save(buffer, format="JPEG", quality=75)
            buffer.seek(0)
            return buffer.getvalue()
    except Exception:
        pass
    return None

def compress_image(uploaded_file):
    image = Image.open(uploaded_file)
    image.thumbnail((1280, 1280))
    buffer = io.BytesIO()
    image.convert("RGB").save(buffer, format="JPEG", quality=80)
    buffer.seek(0)
    return buffer.getvalue()

def upload_to_drive(file_name, file_content, mime_type='image/jpeg'):
    if drive_service is None: return None
    try:
        file_metadata = {'name': file_name, 'parents': [DRIVE_FOLDER_ID]}
        media = MediaIoBaseUpload(io.BytesIO(file_content), mimetype=mime_type, resumable=True)
        file = drive_service.files().create(
            body=file_metadata, media_body=media, fields='id, webViewLink',
            supportsAllDrives=True, supportsTeamDrives=True
        ).execute()
        drive_service.permissions().create(
            fileId=file.get('id'), body={'type': 'anyone', 'role': 'reader'},
            supportsAllDrives=True, supportsTeamDrives=True
        ).execute()
        return file.get('webViewLink')
    except Exception as e:
        st.error(f"드라이브 업로드 오류: {e}")
        return None

def apply_face_blur_ai(img_file):
    try:
        img_file.seek(0)
        compressed_bytes = compress_image(img_file)
        file_bytes = np.asarray(bytearray(compressed_bytes), dtype=np.uint8)
        image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if image is None: return compressed_bytes 
        h, w, _ = image.shape
        pil_img = Image.open(io.BytesIO(compressed_bytes))

        prompt = "이미지에서 모든 사람의 얼굴(머리 전체) 위치를 찾아서 [ymin, xmin, ymax, xmax] 좌표 리스트로 응답해줘. JSON 형식: {\"faces\": [[ymin, xmin, ymax, xmax], ...]}"
        
        response = call_gemini_api_safe(
            client, model_name, [prompt, pil_img],
            config=genai.types.GenerateContentConfig(response_mime_type="application/json")
        )
        face_data = json.loads(response.text)
        faces = face_data.get("faces", [])
        if not faces: return compressed_bytes

        for box in faces:
            ymin, xmin, ymax, xmax = box
            left, top = int(xmin * w / 1000), int(ymin * h / 1000)
            right, bottom = int(xmax * w / 1000), int(ymax * h / 1000)
            
            if top > h * 0.9: # 하단 10% 영역 제외
                continue
                
            rw, rh = right - left, bottom - top
            if rw <= 0 or rh <= 0: continue
            face_roi = image[top:bottom, left:right]
            mask = np.zeros((rh, rw), dtype=np.uint8)
            cv2.circle(mask, (rw // 2, rh // 2), min(rw, rh) // 2, (255), -1)
            blurred_roi = cv2.GaussianBlur(face_roi, (21, 21), 0)
            mask_3ch = cv2.merge([mask, mask, mask])
            image[top:bottom, left:right] = np.where(mask_3ch == 255, blurred_roi, face_roi)

        _, buffer = cv2.imencode('.jpg', image, [cv2.IMWRITE_JPEG_QUALITY, 90])
        return buffer.tobytes()
    except Exception:
        return img_file.getvalue()

def search_kosha_guide(search_keyword):
    decoding_service_key = "801f7d06fa1418ec27119eea23fac9fa6aeec50a1a6e6680ea8197534e50e708"
    endpoint = "https://apis.data.go.kr/B552468/koshaguide/getKoshaGuide"
    params = {'serviceKey': urllib.parse.unquote(decoding_service_key), 'pageNo': '1', 'numOfRows': '5', 'callApiId': '1050', 'techGdlnNm': search_keyword}
    headers = {'User-Agent': 'Mozilla/5.0'}
    try:
        res = requests.get(endpoint, params=params, headers=headers, timeout=5)
        if res.status_code == 200:
            items = res.json().get('body', {}).get('items', {}).get('item', [])
            return [items] if isinstance(items, dict) else items
        return []
    except Exception: return []

def append_row_to_sheet(row_data):
    try:
        body = {'values': [row_data]}
        sheets_service.spreadsheets().values().append(
            spreadsheetId=SPREADSHEET_ID, range=f"'{SHEET_NAME}'!A1",
            valueInputOption='USER_ENTERED', insertDataOption='INSERT_ROWS', body=body
        ).execute()
        return True
    except Exception as e:
        st.error(f"시트 저장 실패: {e}")
        return False

def update_action_result_to_sheet(row_idx, action_data):
    try:
        actual_row = row_idx + 2
        range_name = f"'{SHEET_NAME}'!N{actual_row}:R{actual_row}"
        body = {'values': [action_data]}
        sheets_service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID, range=range_name,
            valueInputOption='USER_ENTERED', body=body
        ).execute()
        return True
    except Exception as e:
        st.error(f"개선조치 업데이트 실패: {e}")
        return False

def load_dashboard_data():
    sheet_url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=csv&gid=413707311"
    try:
        df = pd.read_csv(sheet_url)
        if '타임스탬프' in df.columns:
            df['타임스탬프'] = df['타임스탬프'].astype(str).str.replace('오전', 'AM').str.replace('오후', 'PM')
            df['타임스탬프'] = pd.to_datetime(df['타임스탬프'], format='mixed', errors='coerce')
            df = df.dropna(subset=['타임스탬프'])
        
        needed_cols = ['개선후 빈도', '개선후 강도', '개선후 점수', '개선후 위험등급', '개선후 사진기록']
        for col in needed_cols:
            if col not in df.columns:
                df[col] = np.nan
                
        return df
    except Exception:
        return None

def get_risk_grade(score):
    if score <= 3: return "매우 낮음"
    elif score <= 6: return "낮음"
    elif score <= 9: return "보통"
    elif score <= 12: return "높음"
    else: return "매우 높음"

# =========================================================
# 📥 [내보내기 전용 유틸리티 함수 3종 및 컬럼 상수]
# =========================================================

EXPORT_COLUMNS = [
    '타임스탬프', '시설명', '담당 부서', '장소', '유해위험요인', '위험상황', 
    '빈도', '강도', '점수', '위험등급', '감소대책', '관련근거', '사진 기록', 
    '개선후 빈도', '개선후 강도', '개선후 점수', '개선후 위험등급', '개선후 사진기록'
]

FONT_NAME = 'Helvetica'
font_paths = ['malgun.ttf', 'c:/Windows/Fonts/malgun.ttf']

for fpath in font_paths:
    if os.path.exists(fpath):
        try:
            pdfmetrics.registerFont(TTFont('MalgunGothic', fpath))
            FONT_NAME = 'MalgunGothic'
            break
        except Exception:
            pass

# --- [AI 프롬프트 수정: 서두 문구 금지 및 마크다운 표 지양] ---
PDF_AI_PROMPT_TEMPLATE = """
다음은 {facility_name}의 위험성평가 데이터 통계 및 상세 내역입니다.

[데이터 통계]
- 총 유해위험요인: {total_cnt}건
- 감소대책 수립: {plan_cnt}건 / 개선조치 완료: {complete_cnt}건
- 위험등급 현황: 높음({high_cnt}건), 보통({med_cnt}건), 낮음({low_cnt}건), 매우 낮음({vlow_cnt}건)

[상세 내역]
{risk_details}

[작성 규칙]
1. 절대로 "공공기관 및 공공사업장 보고서 표준 양식..." 같은 인사말이나 서두 안내 문구를 작성하지 마세요.
2. 바로 "# [보고] {facility_name} 사업장 위험성평가 결과 분석 및 개선대책" 제목부터 시작하세요.
3. 마크다운 표(|---|) 형태는 절대로 사용하지 마시고, 개조식 문장(○, -, ☐)으로만 간결하게 작성해 주세요.
4. 항목 제목은 #, ##, ### 및 **강조**를 적절히 활용하여 작성하세요.
"""

# --- [마크다운 -> ReportLab 태그 변환 및 서두 제거 파서] ---
def clean_markdown_for_reportlab(text):
    # 인사말 문구 강제 제거
    text = re.sub(r"^.*?보고서입니다\.\n?", "", text, flags=re.DOTALL)
    
    lines = text.split('\n')
    cleaned_lines = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # 마크다운 표(|) 및 구분선 제거
        if line.startswith('|') or '---|' in line:
            continue
            
        # 마크다운 제목(#)을 Font 크기 태그로 변환
        if line.startswith('# '):
            line = f"<font size=13><b>{line.replace('# ', '').strip()}</b></font>"
        elif line.startswith('## '):
            line = f"<font size=11><b>{line.replace('## ', '').strip()}</b></font>"
        elif line.startswith('### '):
            line = f"<font size=10><b>{line.replace('### ', '').strip()}</b></font>"
            
        # 마크다운 볼드(**) 변환
        line = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', line)
        
        # 특수문자 이스케이프 및 XML 태그 복원
        line = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        line = line.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
        line = line.replace("&lt;font size=13&gt;", "<font size=13>").replace("&lt;font size=11&gt;", "<font size=11>").replace("&lt;font size=10&gt;", "<font size=10>").replace("&lt;/font&gt;", "</font>")
        
        cleaned_lines.append(line)
        
    return cleaned_lines

def generate_ai_summary(export_df, facility_name, client):
    total_cnt = len(export_df)
    plan_cnt = len(export_df[export_df['감소대책'].notna() & (export_df['감소대책'] != '')]) if '감소대책' in export_df.columns else 0
    complete_cnt = len(export_df[export_df['개선후 위험등급'].notna() & (export_df['개선후 위험등급'] != '')]) if '개선후 위험등급' in export_df.columns else 0
    
    high_cnt = len(export_df[export_df['위험등급'] == '높음']) if '위험등급' in export_df.columns else 0
    med_cnt = len(export_df[export_df['위험등급'] == '보통']) if '위험등급' in export_df.columns else 0
    low_cnt = len(export_df[export_df['위험등급'] == '낮음']) if '낮음' in export_df.columns else 0
    vlow_cnt = len(export_df[export_df['위험등급'] == '매우 낮음']) if '매우 낮음' in export_df.columns else 0

    risk_details_text = ""
    if '유해위험요인' in export_df.columns and '감소대책' in export_df.columns:
        risk_details_list = export_df[['유해위험요인', '감소대책']].to_dict(orient='records')
        risk_details_text = "\n".join([f"- 요인: {r.get('유해위험요인', '')} / 대책: {r.get('감소대책', '')}" for r in risk_details_list])

    prompt = PDF_AI_PROMPT_TEMPLATE.format(
        facility_name=facility_name, total_cnt=total_cnt, plan_cnt=plan_cnt,
        complete_cnt=complete_cnt, high_cnt=high_cnt, med_cnt=med_cnt,
        low_cnt=low_cnt, vlow_cnt=vlow_cnt, risk_details=risk_details_text
    )

    try:
        response = call_gemini_api_safe(client, model_name, prompt)
        return response.text
    except Exception:
        return "※ AI 총평 생성 중 일시적인 서버 과부하가 발생하여 통계 데이터 기반 표준 총평으로 대체합니다."

# --- [PDF 생성 함수 수정: A4 -> A3 레이아웃 밀림 방지 및 스타일 최적화] ---
def create_pdf_with_ai_summary(export_df, facility_name, client, export_cols, get_image_bytes_func):
    buffer = io.BytesIO()
    
    # 기본 A4 세로 문서 설정
    doc = SimpleDocTemplate(
        buffer, pagesize=portrait(A4),
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
    )

    frame_a4 = Frame(20, 20, 555.27, 801.89, id='normal_a4')
    template_a4 = PageTemplate(id='A4_Portrait', frames=frame_a4, pagesize=portrait(A4))

    frame_a3 = Frame(20, 20, 1150.55, 801.89, id='normal_a3')
    template_a3 = PageTemplate(id='A3_Landscape', frames=frame_a3, pagesize=landscape(A3))

    doc.addPageTemplates([template_a4, template_a3])

    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontName=FONT_NAME, fontSize=15, leading=18, alignment=1, spaceAfter=10
    )
    body_style = ParagraphStyle(
        'CustomBody', parent=styles['Normal'],
        fontName=FONT_NAME, fontSize=8, leading=11
    )
    table_cell_style = ParagraphStyle(
        'TableCell', parent=styles['Normal'],
        fontName=FONT_NAME, fontSize=7, leading=9, alignment=1
    )

    # --- PAGE 1: 요약 보고서 (A4 세로) ---
    story.append(Paragraph(f"<b>2026년 {facility_name} 위험성평가 결과 보고서</b>", title_style))
    story.append(Spacer(1, 8))

    if client:
        raw_ai_text = generate_ai_summary(export_df, facility_name, client)
        cleaned_lines = clean_markdown_for_reportlab(raw_ai_text)
        
        for line in cleaned_lines:
            story.append(Paragraph(line, body_style))
            story.append(Spacer(1, 2))

    # --- PAGE 2: A3 가로 페이지로 전환 및 첨부 표 배치 ---
    story.append(NextPageTemplate('A3_Landscape'))
    story.append(PageBreak())

    story.append(Paragraph("<b>[첨부] 세부 위험성평가 및 개선조치 현황</b>", title_style))
    story.append(Spacer(1, 5))

    # 1. 헤더 및 표 데이터 구성
    headers = [col for col in export_cols if col in export_df.columns]
    for col in export_df.columns:
        if col not in headers:
            headers.append(col)

    header_paragraphs = [Paragraph(f"<b>{h}</b>", table_cell_style) for h in headers]
    table_data = [header_paragraphs]

    for _, row in export_df.iterrows():
        row_cells = []
        for col_name in headers:
            val = str(row.get(col_name, '')) if pd.notna(row.get(col_name, '')) else ''
            val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            
            if "사진" in str(col_name) and val.startswith("http"):
                row_cells.append(Paragraph("사진 첨부", table_cell_style))
            else:
                row_cells.append(Paragraph(val, table_cell_style))
        table_data.append(row_cells)

    # 2. A3 너비(1150pt) 기준 열 비율 산정
    COLUMN_RATIOS = {
        '타임스탬프': 0.04, '시설명': 0.03, '담당 부서': 0.05, '장소': 0.05, '유해위험요인': 0.05,
        '위험상황': 0.18, '빈도': 0.025, '강도': 0.025, '점수': 0.025, '위험등급': 0.035,
        '감소대책': 0.18, '관련근거': 0.12, '사진 기록': 0.04,
        '개선후 빈도': 0.025, '개선후 강도': 0.025, '개선후 점수': 0.025, '개선후 위험등급': 0.035, '개선후 사진기록': 0.04
    }

    usable_width = 1150.0
    actual_widths = [usable_width * COLUMN_RATIOS.get(h, 0.05) for h in headers]

    # 3. Table 생성 및 패딩 축소 (2페이지 내 밀림 방지)
    pdf_table = Table(table_data, colWidths=actual_widths, repeatRows=1)
    pdf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))

    story.append(pdf_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def create_excel_with_images(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "위험성평가_개선현황"

    headers = [col for col in EXPORT_COLUMNS if col in df.columns]
    for col in df.columns:
        if col not in headers:
            headers.append(col)

    ws.append(headers)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    photo_cols = [i for i, col in enumerate(headers) if "사진" in str(col)]

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[row_idx].height = 80
        for col_idx, col_name in enumerate(headers, start=1):
            value = row.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if pd.notna(value) else "")
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

            if (col_idx - 1) in photo_cols and pd.notna(value) and str(value).startswith("http"):
                img_bytes = get_image_bytes_from_link(str(value))
                if img_bytes:
                    try:
                        img_io = io.BytesIO(img_bytes)
                        img = OpenpyxlImage(img_io)
                        img.width = 100
                        img.height = 75
                        cell_address = f"{get_column_letter(col_idx)}{row_idx}"
                        ws.add_image(img, cell_address)
                        cell.value = ""
                    except Exception:
                        pass

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        col_name = str(headers[col[0].column - 1]).strip()
        
        if "사진" in col_name or col_name in ['타임스탬프', '장소']:
            ws.column_dimensions[col_letter].width = 16.6
        elif col_name == '유해위험요인':
            ws.column_dimensions[col_letter].width = 15.0
        elif col_name == '위험상황':
            ws.column_dimensions[col_letter].width = 40.0
        elif col_name == '감소대책':
            ws.column_dimensions[col_letter].width = 50.0
        elif col_name == '관련근거':
            ws.column_dimensions[col_letter].width = 30.0
        elif col_name == '개선후 위험등급':
            ws.column_dimensions[col_letter].width = 15.0
        else:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def create_hwpx_with_images(df):
    output = io.BytesIO()

    headers = [col for col in EXPORT_COLUMNS if col in df.columns]
    for col in df.columns:
        if col not in headers:
            headers.append(col)

    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('mimetype', 'application/hwp+zip', compress_type=zipfile.ZIP_STORED)

        container_xml = """<?xml version="1.0" encoding="UTF-8"?>
<container xmlns="urn:oasis:names:tc:opendocument:xmlns:container">
    <rootfiles>
        <rootfile full-path="Contents/content.hpf" media-type="application/hwpdoc+xml"/>
    </rootfiles>
</container>"""
        zf.writestr('META-INF/container.xml', container_xml)

        content_hpf = """<?xml version="1.0" encoding="UTF-8"?>
<package xmlns="http://www.hancom.com/hwpml/2011/version" version="1.0">
    <metadata>
        <title>KYWA AI 위험성평가 종합 보고서</title>
    </metadata>
    <manifest>
        <item id="header" href="header.xml" media-type="application/xml"/>
        <item id="section0" href="section0.xml" media-type="application/xml"/>
    </manifest>
    <spine>
        <itemref idref="section0"/>
    </spine>
</package>"""
        zf.writestr('Contents/content.hpf', content_hpf)

        header_xml = """<?xml version="1.0" encoding="UTF-8"?>
<hh:head xmlns:hh="http://www.hancom.com/hwpml/2011/head"/>"""
        zf.writestr('Contents/header.xml', header_xml)

        rows_xml = "<hp:tr>"
        for h in headers:
            rows_xml += f"<hp:tc><hp:p><hp:t>{h}</hp:t></hp:p></hp:tc>"
        rows_xml += "</hp:tr>"

        for _, row in df.iterrows():
            rows_xml += "<hp:tr>"
            for col_name in headers:
                val = str(row.get(col_name, '')) if pd.notna(row.get(col_name, '')) else ''
                val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                rows_xml += f"<hp:tc><hp:p><hp:t>{val}</hp:t></hp:p></hp:tc>"
            rows_xml += "</hp:tr>"

        section_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<hs:sec xmlns:hs="http://www.hancom.com/hwpml/2011/section" xmlns:hp="http://www.hancom.com/hwpml/2011/paragraph">
    <hp:p><hp:t>🚨 KYWA AI 위험성평가 및 개선조치 종합 보고서</hp:t></hp:p>
    <hp:tbl rowCnt="{len(df)+1}" colCnt="{len(headers)}">
        {rows_xml}
    </hp:tbl>
</hs:sec>"""
        zf.writestr('Contents/section0.xml', section_xml)

    output.seek(0)
    return output.getvalue()

# --- [스타일링 및 헤더] ---
st.markdown("""
    <style>
    div.stButton > button {
        background-color: #ff4b4b !important;
        color: white !important;
        font-weight: bold !important;
        border-radius: 0.5rem !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button:hover {
        background-color: #ff3333 !important;
        transform: scale(1.01);
    }
    .logo-container img { height: 48px !important; width: auto !important; }
    </style>
""", unsafe_allow_html=True)

header_col1, header_col2 = st.columns([1, 4])
with header_col1:
    if os.path.exists("kywa_logo.png"):
        with open("kywa_logo.png", "rb") as f:
            logo_base64 = base64.b64encode(f.read()).decode()
        st.markdown(f'''
            <a href="https://www.kywa.or.kr/main/main.jsp" target="_blank" class="logo-container">
                <img src="data:image/png;base64,{logo_base64}" style="height: 48px; margin-top: 5px;">
            </a>
        ''', unsafe_allow_html=True)
    else:
        st.markdown('''<a href="https://www.kywa.or.kr/main/main.jsp" target="_blank">
            <h2 style="color:#ff4b4b; margin-top:10px;">KYWA</h2></a>''', unsafe_allow_html=True)

with header_col2:
    st.markdown("""<h1 style='margin-bottom: 0;'>🚨 KYWA AI 위험성평가 시스템</h1>
    <p style='color: gray; margin-top: 0;'>Korea Youth Work Agency - 스마트 안전관리 플랫폼</p>""", unsafe_allow_html=True)

st.divider()

if "analysis_results" not in st.session_state: st.session_state.analysis_results = None
if "final_data" not in st.session_state: st.session_state.final_data = None
if "eval_after_data" not in st.session_state: st.session_state.eval_after_data = None

# --- [대시보드 출력 함수] ---
def render_dashboard(dashboard_data, key_suffix="default"):
    if dashboard_data is not None:
        if '타임스탬프' in dashboard_data.columns:
            yearly_data = dashboard_data[dashboard_data['타임스탬프'].dt.year == 2026].copy()
        else:
            yearly_data = dashboard_data.copy()

        if yearly_data.empty:
            st.warning("📅 2026년도 데이터가 아직 없습니다. 데이터를 첫 번째로 전송해 보세요!")
        else:
            st.subheader("📊 실시간 점검 데이터 현황 (2026년)")
            total_count = len(yearly_data)
            m1, m2 = st.columns(2)
            
            with m1:
                with st.container(border=True):
                    st.metric("올해 누적 점검 건수", f"{total_count} 건")
            
            with m2:
                with st.container(border=True):
                    author_col = "작성자 성명" 
                    if author_col in yearly_data.columns:
                        st.metric("참여 인원(명)", f"{yearly_data[author_col].nunique()} 명")
                    elif "시설명" in yearly_data.columns:
                        st.metric("점검결과 제출 시설", f"{yearly_data['시설명'].nunique()} 개 시설")
                    else:
                        st.metric("점검결과 제출 건수", f"{total_count} 건")

            CATEGORY_COLOR_MAP = {
                "시설 안전": "#D32F2F", "화재 안전": "#FF5722", "재난 안전": "#880E4F",
                "작업 안전": "#FFA000", "작업 특성": "#E64A19", "기계(설비)적 요인": "#795548",
                "전기적 요인": "#FBC02D", "보건 및 위생관리": "#E91E63", "화학물질 관리": "#9C27B0",
                "작업 환경": "#455A64", "보행 안전": "#1976D2", "활동 안전": "#388E3C"
            }

            FACILITY_COLOR_MAP = {
                "중앙": "#B93444", "본원": "#6B5B95", "평창": "#E2725B",
                "바이오": "#D2B48C", "해양": "#5B84B1", "우주": "#2E4A62",
                "미래": "#92B06A", "생태": "#5F7161"
            }

            g_col1, g_col2 = st.columns(2)

            with g_col1:
                with st.container(border=True):
                    target_col_cat = "위험요인 분류" if "위험요인 분류" in yearly_data.columns else (yearly_data.columns[4] if len(yearly_data.columns) >= 5 else None)
                    if target_col_cat:
                        st.write(f"**⚠️ {target_col_cat} 현황**")
                        if not yearly_data[target_col_cat].dropna().empty:
                            yearly_data[target_col_cat] = yearly_data[target_col_cat].astype(str).str.strip()
                            fig_pie = px.pie(
                                yearly_data, names=target_col_cat, hole=0.3,
                                color=target_col_cat, color_discrete_map=CATEGORY_COLOR_MAP
                            )
                            fig_pie.update_traces(
                                textinfo='percent+value', texttemplate='%{percent:.0%}<br>(%{value}건)',
                                insidetextorientation='horizontal', textfont_size=11
                            )
                            fig_pie.update_layout(
                                margin=dict(t=20, b=60, l=0, r=0), height=400, showlegend=True,
                                legend=dict(orientation="h", yanchor="top", y=-0.1, xanchor="center", x=0.5, font=dict(size=10), itemwidth=30),
                                paper_bgcolor='rgba(0,0,0,0)', dragmode=False
                            )
                            st.plotly_chart(fig_pie, use_container_width=True, config={'displayModeBar': False}, key=f"pie_{key_suffix}")

            with g_col2:
                with st.container(border=True):
                    target_col_fac = "시설명" 
                    if target_col_fac in yearly_data.columns:
                        st.write(f"**🏢 {target_col_fac}별 건수**")
                        yearly_data[target_col_fac] = yearly_data[target_col_fac].astype(str).str.strip()
                        fac_counts = yearly_data[target_col_fac].value_counts().reset_index()
                        fac_counts.columns = [target_col_fac, '건수']
                        
                        fig_bar = px.bar(
                            fac_counts, x=target_col_fac, y='건수', color=target_col_fac,
                            color_discrete_map=FACILITY_COLOR_MAP
                        )
                        fig_bar.update_traces(texttemplate='%{y}건', textposition='outside', textfont_size=11)
                        fig_bar.update_layout(
                            margin=dict(t=20, b=0, l=0, r=0), height=400, showlegend=False,
                            xaxis_title=None, yaxis_title=None, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', dragmode=False 
                        )
                        st.plotly_chart(fig_bar, use_container_width=True, config={'displayModeBar': False}, key=f"bar_{key_suffix}")

# --- [사이드바 메뉴] ---
with st.sidebar:
    st.markdown("### 📌 메뉴 선택")
    selected_tab = st.radio(
        label="이동할 화면을 선택하세요",
        options=["📝 점검 입력(전 직원)", "📊 개선조치 및 재평가(담당자)", "📥 결과보고서(관리자)"],
        label_visibility="collapsed"
    )

# ==========================================
# 1️⃣ [메뉴 1] 점검 입력
# ==========================================
if selected_tab == "📝 점검 입력(전 직원)":
    col1, col2 = st.columns([1, 1])
    with col1:
        st.markdown("### **🏢 점검 대상 정보**")
        selected_facility = st.radio("• 시설명 선택 (필수)", ["중앙", "평창", "우주", "바이오", "해양", "미래", "생태", "본원"], horizontal=True, key="t1_fac")
        dept_list = ["협력부(국립청소년시설)", "활동부(국립청소년시설)", "근로자대표", "청소년성장지원부", "지도인력양성부", "지도인력개발부", "청렴감사실", "기획혁신부", "인재경영부", "홍보전략부", "안전경영부", "재무회계부", "디지털정보부", "활동기획부", "미래활동부", "정책사업부", "활동안전부", "활동인증부", "자회사", "협력업체(공사, 용역 등)"]
        selected_dept = st.selectbox("• 담당 부서 선택 (필수)", dept_list, key="t1_dept")
        st.markdown("### **📝 현장 상황 설명**")
        user_description = st.text_area("• 상황 설명 입력 (권장)", placeholder="1. 본관 2층 테라스 난간 흔들림\n2. 정문 보도블록 파손", height=150, key="t1_desc")

    with col2:
        st.markdown("### **📸 사진 기록 방식**")
        source_option = st.radio("사진 방식 선택", ("📸 사진", "🚫 없음"), horizontal=True, label_visibility="collapsed", key="t1_src")
        img_file = None
        if "📸" in source_option:
            st.info("📸 아래 박스를 클릭하면 [사진촬영] 또는 [사진업로드] 선택이 가능합니다. ✅가로사진 권장")
            img_file = st.file_uploader("사진 업로드", type=['png', 'jpg', 'jpeg'], label_visibility="collapsed", key="t1_img")
            if img_file:
                st.image(img_file, caption="업로드된 원본 사진", use_container_width=True)

    if st.button("🚀 KYWA AI 위험요인 분석 시작", use_container_width=True, key="btn_run_analysis"):
        if not user_description.strip() and not img_file:
            st.warning("⚠️ 분석할 내용(글 또는 사진)을 입력해 주세요.")
        else:
            with st.spinner(f"✨ KYWA AI가 [{selected_facility}] 시설 데이터를 분석 중입니다..."):
                content = []
                prompt = f"""
                당신은 한국청소년활동진흥원(KYWA) 안전관리 전문가입니다.
                [시설: {selected_facility}, 부서: {selected_dept}, 상황: {user_description}]
                위험성평가를 수행하여 JSON 리스트로 출력하세요.
                키: category, location, scenario, p, s, score, grade, law, solution
                """
                content.append(prompt)
                if img_file:
                    processed_bytes = apply_face_blur_ai(img_file)
                    content.append(Image.open(io.BytesIO(processed_bytes)))
                    st.session_state.final_secure_image = processed_bytes

                try:
                    response = call_gemini_api_safe(
                        client, model_name, content,
                        config={"response_mime_type": "application/json", "temperature": 0.0}
                    )
                    if response:
                        res_data = json.loads(response.text.strip())
                        st.session_state.analysis_results = res_data if isinstance(res_data, list) else [res_data]
                        st.success("✅ 분석 완료!")
                        st.rerun()
                except Exception as e:
                    st.error(f"❌ 분석 실패 (서버 트래픽 과부하 등): {e}")

    # --- [AI 위험성평가 결과 편집 및 KOSHA GUIDE 조회] ---
    if st.session_state.analysis_results:
        st.markdown("### 📋 AI 위험성평가 결과")
        df = pd.DataFrame(st.session_state.analysis_results)
        
        # 전송 여부 선택 컬럼(기본값 True) 추가
        if "send_check" not in df.columns:
            df.insert(0, "send_check", True)

        edited_df = st.data_editor(
            df,
            column_config={
                "send_check": st.column_config.CheckboxColumn("전송선택", default=True, width="small"),
                "category": st.column_config.TextColumn("분류", disabled=True),
                "location": st.column_config.TextColumn("📍장소(편집가능)", width="medium"),
                "scenario": st.column_config.TextColumn("✅ 위험상황(편집가능)", width="medium"),
                "p": st.column_config.TextColumn("빈도", disabled=True, width="small"),
                "s": st.column_config.TextColumn("강도", disabled=True, width="small"),
                "score": st.column_config.TextColumn("점수", disabled=True, width="small"),
                "grade": st.column_config.TextColumn("등급", disabled=True, width="small"),
                "law": st.column_config.TextColumn("관련근거", disabled=True, width="medium"),
                "solution": st.column_config.TextColumn("✅ 감소대책(편집가능)", width="large")
            },
            disabled=["category", "p", "s", "score", "grade", "law"],
            hide_index=True, key="t1_editor"
        )
        
        # 전체 결과 저장은 유지
        st.session_state.final_data = edited_df.to_dict('records')
        
        # 전송 체크박스가 선택된 항목만 별도로 추출하여 저장
        selected_records = edited_df[edited_df["send_check"] == True].to_dict('records')
        st.session_state.selected_data_to_send = selected_records

        st.write("")
        with st.expander("📚 **관련 KOSHA GUIDE (자율 안전보건가이드) 조회 및 다운로드**", expanded=True):
            if edited_df is not None and not edited_df.empty:
                search_kw = "추락"
                first_row = edited_df.iloc[0]
                full_text = f"{first_row.get('scenario','')} {first_row.get('law','')}"
                for kw in ['용접', '비계', '사다리', '지게차', '크레인', '개구부', '난간', '추락', '감전', '화재']:
                    if kw in full_text:
                        search_kw = kw
                        break
                guides = search_kosha_guide(search_kw)
                if guides:
                    st.success(f"키워드 **'{search_kw}'** 관련 코샤가이드 {len(guides)}건이 검색되었습니다.")
                    for g in guides:
                        c1, c2 = st.columns([3, 1])
                        c1.markdown(f"• **[{g.get('techGdlnNo','')}]** {g.get('techGdlnNm','')}")
                        if g.get('fileDownloadUrl'):
                            c2.link_button("📥 지침 다운로드", g.get('fileDownloadUrl'), use_container_width=True)
                else:
                    st.info(f"키워드 '{search_kw}' 검색 결과가 없습니다.")
            else:
                st.info("💡 위험성평가 분석을 실행하시면, 맞춤형 KOSHA GUIDE 원문 다운로드 링크가 제공됩니다.")

        st.write("")
        if st.button("✅ KYWA AI 안전센터로 데이터 최종 전송", use_container_width=True, key="btn_send_t1"):
            if sheets_service is None:
                st.error("⚠️ GCP 서비스 연동이 필요합니다.")
            elif not st.session_state.get("final_data"):
                st.error("⚠️ 전송할 데이터가 없습니다.")
            else:
                with st.spinner("🚀 KYWA AI 안전센터로 데이터 전송 중..."):
                    now_kst = datetime.datetime.now() + datetime.timedelta(hours=9)
                    current_time = now_kst.strftime("%Y-%m-%d %H:%M:%S")
                    timestamp_str = now_kst.strftime("%Y%m%d_%H%M%S")
                    photo_link = "사진 없음"
                    if "final_secure_image" in st.session_state and st.session_state.final_secure_image:
                        photo_link = upload_to_drive(f"{selected_facility}_{timestamp_str}.jpg", st.session_state.final_secure_image)

                    # 선택된 데이터 가져오기 (체크박스로 선택된 항목만 필터링)
                    target_data = st.session_state.get("selected_data_to_send", st.session_state.final_data)

                    success_count = 0
                    for row in target_data:
                        sheet_row = [
                            current_time, selected_facility, selected_dept,
                            row.get("location"), row.get("category"), row.get("scenario"),
                            row.get("p"), row.get("s"), row.get("score"), row.get("grade"),
                            row.get("solution"), row.get("law"), photo_link
                        ]
                        if append_row_to_sheet(sheet_row):
                            success_count += 1

                    if success_count > 0:
                        st.success(f"✅ {success_count}건의 데이터가 성공적으로 전송되었습니다!")
                        st.balloons()
    st.write("---")
    dashboard_data = load_dashboard_data()
    render_dashboard(dashboard_data, key_suffix="tab1")

# ==========================================
# 2️⃣ [메뉴 2] 결과 조회 및 개선조치
# ==========================================
elif selected_tab == "📊 개선조치 및 재평가(담당자)":
    st.markdown("### 🛠️ 현장 개선조치 결과 등록")
    
    # 담당자 비밀번호 검증
    staff_pw = st.text_input("🔑 담당자 비밀번호 입력", type="password", key="t2_pw")
    if staff_pw == "1234":  # 원하시는 담당자 비밀번호 설정
        st.success("🔓 담당자 인증 성공")
        st.info("시설을 선택하고 '미완료' 건을 조회하여 조치 결과 사진 및 내용을 입력하세요.")

        dashboard_data = load_dashboard_data()
        if dashboard_data is None or dashboard_data.empty:
            st.warning("⚠️ 불러올 점검 데이터가 없습니다.")
        else:
            target_fac = st.selectbox("• 시설명 선택", ["중앙", "평창", "우주", "바이오", "해양", "미래", "생태", "본원"], key="t2_fac")
            fac_df = dashboard_data[dashboard_data['시설명'] == target_fac].copy()

            uncompleted_df = fac_df[
                fac_df['개선후 위험등급'].isna() | 
                (fac_df['개선후 위험등급'].astype(str).str.strip() == '') | 
                (fac_df['개선후 위험등급'].astype(str).str.strip() == 'nan')
            ]

            if uncompleted_df.empty:
                st.success(f"🎉 [{target_fac}] 시설은 개선 조치가 필요한 항목이 없습니다.")
            else:
                options = {}
                for idx, row in uncompleted_df.iterrows():
                    label = f"[{row.get('타임스탬프', '일시미상')}] {row.get('장소', '장소미상')} - {str(row.get('위험상황', ''))[:20]}..."
                    options[label] = idx

                selected_label = st.selectbox("• 미완료 위험성평가 건 선택", list(options.keys()))
                selected_row_idx = options[selected_label]
                target_item = uncompleted_df.loc[selected_row_idx]

                st.divider()
                st.markdown("#### 📌 선택된 평가건 상세 내용 (개선 전)")
                c1, c2, c3 = st.columns(3)
                c1.write(f"**📍 장소:** {target_item.get('장소', '-')}")
                c2.write(f"**⚠️ 위험상황:** {target_item.get('위험상황', '-')}")
                c3.write(f"**🚨 기존 위험등급:** {target_item.get('위험등급', '-')} ({target_item.get('점수', '-')}점)")

                st.divider()
                st.markdown("#### 📝 개선 후 조치사항 입력")
                
                act_col1, act_col2 = st.columns(2)
                with act_col1:
                    action_text = st.text_area("• 개선조치 상세 내용 입력", placeholder="예: 난간 보강 조치 완료 및 이중 안전고리 설치", height=120)
                with act_col2:
                    action_img = st.file_uploader("📸 개선 후 사진 업로드", type=['png', 'jpg', 'jpeg'], key="t2_act_img")
                    if action_img: st.image(action_img, caption="개선 후 사진", width=250)

                if st.button("🛠️ 조치결과 AI 분석 (빈도/강도 재산출)", use_container_width=True, key="btn_eval_after"):
                    if not action_text.strip():
                        st.warning("⚠️ 개선 조치 상세 내용을 입력해야 AI 분석이 가능합니다.")
                    else:
                        with st.spinner("✨ KYWA AI가 개선 후 위험도를 재평가 중입니다..."):
                            prompt_eval = f"""
                            당신은 안전관리 전문가입니다.
                            [기존 위험상황: {target_item.get('위험상황')}]
                            [기존 위험등급: {target_item.get('위험등급')}, 점수: {target_item.get('점수')}]
                            [개선 조치 내용: {action_text}]

                            위 조치로 인해 감소된 개선 후의 위험도를 평가하세요.
                            빈도(p_after: 1~5), 강도(s_after: 1~4)를 정하고, 점수(score_after = p * s), 위험등급(grade_after)을 산출하세요.
                            JSON 형식: {{"p_after": 1, "s_after": 1, "score_after": 1, "grade_after": "매우 낮음"}}
                            """
                            try:
                                eval_res = call_gemini_api_safe(
                                    client, model_name, [prompt_eval],
                                    config={"response_mime_type": "application/json"}
                                )
                                st.session_state.eval_after_data = json.loads(eval_res.text.strip())
                                st.success("✅ 개선 후 위험도 재산출 완료!")
                            except Exception as e:
                                st.error(f"❌ 재분석 실패: {e}")

        if st.session_state.eval_after_data:
            res_a = st.session_state.eval_after_data
            st.markdown("##### 📊 KYWA AI 개선 후 위험도 재산출 결과(+ - 클릭으로 변경 가능)")
            
            if "p_after_val" not in st.session_state:
                st.session_state.p_after_val = int(res_a.get("p_after", 1))
            if "s_after_val" not in st.session_state:
                st.session_state.s_after_val = int(res_a.get("s_after", 1))

            r_col1, r_col2, r_col3, r_col4 = st.columns(4)
            p_after = r_col1.number_input("개선후 빈도", min_value=1, max_value=5, value=st.session_state.p_after_val, key="num_p_after")
            s_after = r_col2.number_input("개선후 강도", min_value=1, max_value=4, value=st.session_state.s_after_val, key="num_s_after")
            
            score_after = p_after * s_after
            grade_after = get_risk_grade(score_after)
            
            r_col3.metric("개선후 점수", f"{score_after} 점")
            r_col4.metric("개선후 위험등급", grade_after)

            st.write("")
            if st.button("✅ KYWA AI 안전센터로 데이터 최종 전송", use_container_width=True, key="btn_save_action"):
                with st.spinner("🚀 데이터 전송 및 조치결과를 반영 중입니다..."):
                    act_photo_link = "사진 없음"
                    if 'action_img' in locals() and action_img:
                        act_photo_bytes = apply_face_blur_ai(action_img)
                        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        act_photo_link = upload_to_drive(f"AFTER_{target_fac}_{now_str}.jpg", act_photo_bytes)

                    action_row_data = [
                        p_after, s_after, score_after, grade_after, act_photo_link
                    ]

                    if update_action_result_to_sheet(selected_row_idx, action_row_data):
                        st.success(f"✅ 조치결과(점수: {score_after}점 / 등급: {grade_after})가 KYWA AI 안전센터 DB에 반영되었습니다!")
                        st.session_state.eval_after_data = None
                        if "p_after_val" in st.session_state: del st.session_state.p_after_val
                        if "s_after_val" in st.session_state: del st.session_state.s_after_val
                        st.balloons()

        st.write("---")
        render_dashboard(dashboard_data, key_suffix="tab2")

    elif staff_pw:
        st.error("❌ 비밀번호가 올바르지 않습니다.")

# ==========================================
# 3️⃣ [메뉴 3] 내보내기 (관리자용)
# ==========================================
elif selected_tab == "📥 결과보고서(관리자)":
    st.markdown("### 📊 관리자 전용 종합 대시보드 및 보고서 출력")
    admin_pw = st.text_input("🔑 관리자 비밀번호 입력", type="password")
    
    if admin_pw == "1234":
        st.success("🔓 관리자 인증 성공")
        dashboard_data = load_dashboard_data()
        
        if dashboard_data is not None and not dashboard_data.empty:
            render_dashboard(dashboard_data, key_suffix="tab3")
            
            st.divider()
            st.markdown("#### 📂 전체 위험성평가 및 개선 현황 DB")
            st.dataframe(dashboard_data, use_container_width=True)

            st.divider()
            st.markdown("#### 📥 보고서 및 데이터 맞춤 내보내기")
            
            with st.container(border=True):
                st.markdown("##### **[내보내기 설정]**")
                exp_col1, exp_col2 = st.columns(2)
                
                with exp_col1:
                    export_type = st.radio(
                        "1️⃣ 내보낼 파일 형식 선택",
                        ["📊 엑셀 문서 (.xlsx)", "📄 한글 문서 (.hwpx)", "📄 PDF 보고서 (.pdf)"],
                        key="exp_type"
                    )
                
                facility_options = ["전체"] + ["중앙", "평창", "우주", "바이오", "해양", "미래", "생태", "본원"]
                with exp_col2:
                    selected_exp_fac = st.selectbox(
                        "2️⃣ 대상 시설 선택",
                        facility_options,
                        key="exp_fac"
                    )

                st.write("")
                
                if st.button("🚀 선택한 조건으로 문서 생성하기", use_container_width=True, type="primary", key="btn_generate_doc"):
                    if selected_exp_fac == "전체":
                        export_df = dashboard_data.copy()
                    else:
                        export_df = dashboard_data[dashboard_data['시설명'] == selected_exp_fac].copy()

                    if not export_df.empty:
                        fac_name_str = "전체시설" if selected_exp_fac == "전체" else selected_exp_fac
                        today_str = datetime.date.today().strftime("%Y%m%d")

                        if "엑셀" in export_type:
                            with st.spinner("📊 엑셀 이미지 결합 및 파일 생성 중..."):
                                file_bytes = create_excel_with_images(export_df)
                                file_name = f"KYWA_위험성평가_보고서_{fac_name_str}_{today_str}.xlsx"
                                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        elif "한글" in export_type:
                            with st.spinner("📄 한글(.hwpx) 표 및 데이터 생성 중..."):
                                file_bytes = create_hwpx_with_images(export_df)
                                file_name = f"KYWA_위험성평가_보고서_{fac_name_str}_{today_str}.hwpx"
                                mime_type = "application/hwp+zip"
                        else: # PDF 보고서
                            with st.spinner("📄 AI 총평 포함 PDF 보고서 생성 중..."):
                                file_bytes = create_pdf_with_ai_summary(
                                    export_df=export_df,
                                    facility_name=fac_name_str,
                                    client=client,
                                    export_cols=EXPORT_COLUMNS,
                                    get_image_bytes_func=get_image_bytes_from_link
                                )
                                file_name = f"KYWA_위험성평가_보고서_{fac_name_str}_{today_str}.pdf"
                                mime_type = "application/pdf"

                        st.session_state.export_file_bytes = file_bytes
                        st.session_state.export_file_name = file_name
                        st.session_state.export_file_mime = mime_type
                        st.success(f"✅ [{fac_name_str}] 문서 생성이 완료되었습니다! 아래 다운로드 버튼을 눌러주세요.")
                    else:
                        st.warning("⚠️ 선택한 시설의 점검 데이터가 없습니다.")

                if "export_file_bytes" in st.session_state and st.session_state.export_file_bytes:
                    st.write("")
                    st.download_button(
                        label=f"💾 {st.session_state.export_file_name} 다운로드",
                        data=st.session_state.export_file_bytes,
                        file_name=st.session_state.export_file_name,
                        mime=st.session_state.export_file_mime,
                        use_container_width=True
                    )

    elif admin_pw:
        st.error("❌ 비밀번호가 올바르지 않습니다. 필요시 관리자에게 문의 바랍니다.")

# --- [푸터(Footer) 섹션] ---
st.write("") 
st.write("---")
footer_cols = st.columns([3, 1])

with footer_cols[0]:
    st.markdown("### 🔒 Data Governance & Privacy")
    st.caption("""
    **© 2026 한국청소년활동진흥원(KYWA) 안전경영부.** 본 시스템은 **공공기관 AI 활용 가이드라인** 및 **정보보안 업무규정** 을 엄격히 준수합니다.
    
    * **데이터 보안:** 입력된 모든 정보는 **API 옵트아웃(Opt-out) 설정**이 적용되어 외부 모델 학습에 활용되지 않습니다.
    * **운영 방침:** **KYWA AI 안전센터**로 전송된 데이터는 **담당자의 데이터 적합성 검토**를 거칩니다. 
      점검 내용이 부적절하거나 중복된 경우, 데이터 신뢰성 유지를 위해 운영 관리자에 의해 임의 수정 또는 삭제될 수 있습니다.
    * **면책 고지:** AI 분석 정보는 위험 요인 발굴을 돕는 가이드라인입니다. 실제 위험성 평가 시에는 현장 상황을 반영한 담당 직원의 면밀한 검토를 권고합니다.
    """)

with footer_cols[1]:
    st.markdown("### 📞 Contact")
    st.markdown("""
    <div style="line-height: 1.6;">
        <span style="font-weight: bold; font-size: 0.9rem; color: #31333F;">경영지원본부 안전경영부</span><br>
        <span style="color: #444; font-size: 0.85rem;">📧 archi01@kywa.or.kr</span><br>
        <span style="color: #444; font-size: 0.85rem;">
            <span style="display: inline-block; transform: rotate(10deg); color: #000;">📞</span> 02-6959-7138
        </span>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<p style='font-size: 0.8rem; color: gray; text-align: center;'>Safe Together, KYWA AI Risk Assessment System</p>", unsafe_allow_html=True)
